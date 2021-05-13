import os
import datetime

import xlrd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side

from common.file_operation import create_purchase_output_directory
from storage.models import ProductInfo
from syc import settings

alignment_center = Alignment(horizontal='center',
                             vertical='center',
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)

bd = Border(left=Side(border_style="thin",
                      color='FF001000'),
            right=Side(border_style="thin",
                       color='FF110000'),
            top=Side(border_style="thin",
                     color='FF110000'),
            bottom=Side(border_style="thin",
                        color='FF110000'))


# 全部单元格居中, 并加边框
def align_to_center_border(ws):
    for row in ws.rows:
        for cell in row:
            cell.alignment = alignment_center
            cell.border = bd


def excel_to_pro(path, column_num, sheet_num):
    data = xlrd.open_workbook(path)
    table = data.sheets()[sheet_num]
    check = table.row_values(0)[0]
    lt = [table.row_values(i)[:column_num] for i in range(2, table.nrows - 1)]
    return lt, check


def export_to_excel(data):
    time_now = datetime.datetime.strftime(datetime.datetime.now(), '%Y-%m-%d')
    file_path = os.path.join(settings.BASE_DIR, 'file_bak' + os.sep + 'plan' + os.sep + time_now + '.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.title = '生产安排'
    ws.cell(row=1, column=1, value=time_now + '生产安排')
    for i in range(1, 6):
        col_list = ['商品编号', '数量', '半成品需求量', '成品需求量', '备注']
        ws.cell(row=2, column=i, value=col_list[i - 1])
    for i in range(3, len(data) + 3):
        for j in range(1, 6):
            ws.cell(row=i, column=j, value=data[i - 3][j - 1])
        align_to_center_border(ws)
    wb.save(file_path)
    return file_path


def translate_purchase_excel(path):
    data = xlrd.open_workbook(path)
    table = data.sheet_by_name('Sheet1')
    check = table.row_values(0)[0]
    lt = [(table.row_values(i)[1], int(table.row_values(i)[3])) for i in range(7, table.nrows - 3)]
    return lt


def export_translation_excel(data, file_name):
    file_path = os.path.join(settings.PURCHASE_TRANSLATION_ORIGIN, file_name + '.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=file_name)
    for i in range(1, 3):
        col_list = ['商品编号', '数量']
        ws.cell(row=2, column=i, value=col_list[i - 1])
    for i in range(3, len(data) + 3):
        for j in range(1, 3):
            ws.cell(row=i, column=j, value=data[i - 3][j - 1])
    ws.cell(row=len(data) + 3, column=1, value='总计:')
    align_to_center_border(ws)
    wb.save(file_path)
    return file_path


def export_purchase_list(data, file_name):
    path = create_purchase_output_directory()
    month = str(datetime.datetime.now().month)
    day = str(datetime.datetime.now().day)
    file_path = os.path.join(path, month + '.' + day + '_' + file_name + '.xlsx')
    wb = Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value=file_name)
    row = 3
    total_price = 0
    for i in range(1, 4):
        col_list = ['商品编号', '数量', '标准零售价']
        ws.cell(row=2, column=i, value=col_list[i - 1])
    for i in data:
        total_price += i[1]
        column = 1
        for j in i:
            if column == 1:
                ws.cell(row=row, column=column, value=j)
            elif column == 2:
                ws.cell(row=row, column=column, value=int(j))
            elif column == 3:
                ws.cell(row=row, column=column, value=round(j, 2)).number_format = '0.00'
            column += 1
        row += 1
    ws.cell(row=row, column=1, value='总计')
    ws.cell(row=row, column=2, value=total_price)
    align_to_center_border(ws)
    wb.save(file_path)
    return file_name


def create_statement(cust_info, storage_out_info, path, total_count, total_price, total_actual_price, statement_data):
    date = storage_out_info.storage_create_date.date().strftime('%m月%d日')
    storage_out_detail = storage_out_info.storage_out_detail
    storage_product_info = ProductInfo.objects.all()
    row = 3
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('C2:D2')
    ws.merge_cells('F2:G2')

    # 表头
    for i in range(1, 9):
        col_list = ['', '', '', cust_info.cust_location + cust_info.cust_name, '发货单（）', '',
                    date, '单清A']
        ws.cell(row=1, column=i, value=col_list[i - 1])
    for i in range(1, 9):
        col_list = ['编号', '数量', '零售标准价', '', '金额', '备注', '', '']
        ws.cell(row=2, column=i, value=col_list[i - 1])

    # 发货信息
    for i in storage_out_detail:
        for j in range(1, 6):
            col_list = [i.pro_id, i.storage_pro_count, storage_product_info.get(pro_id=i.pro_id).pro_unit_price, '',
                        i.storage_pro_price]
            if j == 3 or j == 5:
                ws.cell(row=row, column=j, value=col_list[j - 1]).number_format = '0.00'
            else:
                ws.cell(row=row, column=j, value=col_list[j - 1])
        row += 1

    # 其他信息
    for i in range(1, 6):
        col_list = ['小计', total_count, '', '', total_price]
        ws.cell(row=row, column=i, value=col_list[i - 1])
    row += 1
    for i in range(1, 6):
        col_list = ['提货金额', '', '', '', total_actual_price]
        ws.cell(row=row, column=i, value=col_list[i - 1])
    row += 1

    # 历史信息表头
    for i in range(1, 9):
        col_list = ['', '', '对账单', '', '到货请点总数再分类清点', '', '', '']
        ws.cell(row=row, column=i, value=col_list[i - 1])
    row += 1
    for i in range(1, 9):
        col_list = ['日期', '原余款', '收货款', '发货金额', '换货金额', '运费', '余额', '']
        ws.cell(row=row, column=i, value=col_list[i - 1])
    row += 1

    # 历史信息明细
    for i in statement_data:
        # create_date, origin_balance, income, total_actual_price, refund_price, translation_expense, balance
        ws.cell(row=row, column=1, value=i.create_date)
        ws.cell(row=row, column=2, value=i.origin_balance)
        ws.cell(row=row, column=3, value=i.income)
        ws.cell(row=row, column=4, value=i.total_actual_price)
        ws.cell(row=row, column=5, value=i.refund_price)
        ws.cell(row=row, column=6, value=i.translation_expense)
        ws.cell(row=row, column=7, value=i.balance)
        row += 1

    month = storage_out_info.storage_create_date.strftime('%m')
    day = storage_out_info.storage_create_date.strftime('%d')
    file_name = month + '.' + day + '_' + cust_info.cust_location + cust_info.cust_name
    file_path = os.path.join(path, file_name + '.xlsx')
    align_to_center_border(ws)
    wb.save(file_path)
    return file_path, file_name


def create_purchase_statement(cust_info, purchase_info, path, total_count, total_price, total_actual_price,
                              statement_data):
    date = purchase_info.pur_modify_date.strftime('%m月%d日')
    purchase_detail = purchase_info.purchase_detail
    storage_product_info = ProductInfo.objects.all()
    row = 3
    wb = Workbook()
    ws = wb.active
    ws.merge_cells('C2:D2')
    ws.merge_cells('F2:G2')

    # 表头
    for i in range(1, 9):
        col_list = ['', '', '', cust_info.cust_location + cust_info.cust_name, '发货单（）', '',
                    date, '单清A']
        ws.cell(row=1, column=i, value=col_list[i - 1])
    for i in range(1, 9):
        col_list = ['编号', '数量', '零售标准价', '', '金额', '备注', '', '']
        ws.cell(row=2, column=i, value=col_list[i - 1])

    # 发货信息
    for i in purchase_detail:
        for j in range(1, 6):
            col_list = [i.pro_id, i.pur_pro_count, storage_product_info.get(pro_id=i.pro_id).pro_unit_price, '',
                        i.pur_pro_price]
            if j == 3 or j == 5:
                ws.cell(row=row, column=j, value=col_list[j - 1]).number_format = '0.00'
            else:
                ws.cell(row=row, column=j, value=col_list[j - 1])
        row += 1

    # 其他信息
    for i in range(1, 6):
        col_list = ['小计', total_count, '', '', total_price]
        ws.cell(row=row, column=i, value=col_list[i - 1])
    row += 1
    for i in range(1, 6):
        col_list = ['提货金额', '', '', '', total_actual_price]
        ws.cell(row=row, column=i, value=col_list[i - 1])
    row += 1

    # 历史信息表头
    for i in range(1, 9):
        col_list = ['', '', '对账单', '', '到货请点总数再分类清点', '', '', '']
        ws.cell(row=row, column=i, value=col_list[i - 1])
    row += 1
    for i in range(1, 9):
        col_list = ['日期', '原余款', '收货款', '发货金额', '换货金额', '运费', '余额', '']
        ws.cell(row=row, column=i, value=col_list[i - 1])
    row += 1

    # 历史信息明细
    for i in statement_data:
        # create_date, origin_balance, income, total_actual_price, refund_price, translation_expense, balance
        ws.cell(row=row, column=1, value=i.create_date)
        ws.cell(row=row, column=2, value=i.origin_balance)
        ws.cell(row=row, column=3, value=i.income)
        ws.cell(row=row, column=4, value=i.total_actual_price)
        ws.cell(row=row, column=5, value=i.refund_price)
        ws.cell(row=row, column=6, value=i.translation_expense)
        ws.cell(row=row, column=7, value=i.balance)
        row += 1

    month = purchase_info.pur_modify_date.strftime('%m')
    day = purchase_info.pur_modify_date.strftime('%d')
    file_name = month + '.' + day + '_' + cust_info.cust_location + cust_info.cust_name
    file_path = os.path.join(path, file_name + '.xlsx')
    align_to_center_border(ws)
    wb.save(file_path)
    return file_path, file_name
